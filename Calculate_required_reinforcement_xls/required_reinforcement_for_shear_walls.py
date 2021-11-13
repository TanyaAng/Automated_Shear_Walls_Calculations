import openpyxl
import numpy as np
import pandas as pd
from math import ceil

workbook = openpyxl.load_workbook('Shear walls_TOWER.xlsx')
sheet = workbook['Shear Walls']


REBARS_DIAMETER = [8, 10, 12, 14, 16, 18, 20, 22, 25, 28, 32]
REBARS_CM2 = [0.503, 0.785, 1.131, 1.539, 2.011, 2.545, 3.142, 3.801, 4.909, 6.158]
REBAR_WEIGHT = [0.395, 0.617, 0.888, 1.208, 1.998, 2.466, 2.984, 3.853, 4.834, 6.313, 7.990, 9.865]

# not in use
def find_starting_indices_of_required_reinfocement(row_indices, column_indices):
    starting_indices = []
    for x_indices in range(row_indices, 15):
        for y_indices in range(column_indices, 15):
            if sheet.cell(row=x_indices, column=y_indices).value == 'Aa1':
                starting_indices.append(x_indices)
                starting_indices.append(y_indices + 1)
                break
    return starting_indices

# not in use
def find_required_reinforcement(row_indices, column_indices):
    requered_reinforcement = []
    for index in range(row_indices, row_indices + 4):
        requered_reinforcement.append(sheet.cell(row=index, column=column_indices).value)
    return requered_reinforcement


def find_head_size(row_indices, column_indices):
    column_size = []
    for x_indices in range(row_indices, row_indices + 15):
        for y_indices in range(column_indices, column_indices + 9):
            if sheet.cell(row=x_indices, column=y_indices).value == 'глава':
                column_size.append(sheet.cell(row=x_indices, column=y_indices + 4).value)
                column_size.append(sheet.cell(row=x_indices, column=y_indices + 5).value)
                break
    return column_size


def calculate_rebars_in_head(height, width):
    number_rebars = 0
    number_rebars += ceil((height - 5) / 15 + 1) * 2
    if width > 25:
        number_rebars += ceil((width - 5) / 15 - 1) * 2
    return number_rebars


def calculate_Aa1_Aa2(required_reinforcement, number_of_bars):
    required_diameter = required_reinforcement / number_of_bars
    for index in range(len(REBARS_CM2)):
        if required_diameter <= REBARS_CM2[index]:
            provided_diameter = REBARS_DIAMETER[index]
            break
        else:
            provided_diameter = 'Change column size!!!'
    return provided_diameter


def calculate_AaH(required_reinforcement):
    steps = [20, 15, 12.5, 10]
    rebars_per_meter = [round(100 / x, 2) for x in steps]
    min_weight = 10000000
    index_of_min_rebar = 0
    index_of_min_step = 0
    for rebar in REBARS_CM2:
        needed_rebars_per_meter = required_reinforcement / rebar
        index = REBARS_CM2.index(rebar)
        current_rebar_weight = REBAR_WEIGHT[index]
        for i in range(len(rebars_per_meter)):
            if needed_rebars_per_meter > rebars_per_meter[i]:
                continue
            else:
                weight_of_current_mesh = rebars_per_meter[i] * current_rebar_weight
            if 0 < weight_of_current_mesh <= min_weight:
                min_weight = weight_of_current_mesh
                index_of_min_rebar = index
                index_of_min_step = i
    return rebars_per_meter[index_of_min_step], REBARS_DIAMETER[index_of_min_rebar], steps[index_of_min_step]


def calculate_AaV(required_reinforcement):
    steps = [20, 15, 10]
    rebars_per_meter = [round(100 / x, 2) for x in steps]
    min_weight = 10000000
    index_of_min_rebar = 0
    index_of_min_step = 0
    for rebar in REBARS_CM2:
        needed_rebars_per_meter = required_reinforcement / rebar
        index = REBARS_CM2.index(rebar)
        current_rebar_weight = REBAR_WEIGHT[index]
        for i in range(len(rebars_per_meter)):
            if needed_rebars_per_meter > rebars_per_meter[i]:
                continue
            else:
                weight_of_current_mesh = rebars_per_meter[i] * current_rebar_weight
            if 0 < weight_of_current_mesh <= min_weight:
                min_weight = weight_of_current_mesh
                index_of_min_rebar = index
                index_of_min_step = i
    return rebars_per_meter[index_of_min_step], REBARS_DIAMETER[index_of_min_rebar], steps[index_of_min_step]


def find_required_index(starting_row, step):
    row_alpha = []
    row_digit = []
    for index in starting_row:
        if index.isalpha():
            row_alpha.append(index)
        elif index.isdigit():
            row_digit.append(index)
    row_alpha = list(reversed(row_alpha))
    next_row_alpha = []
    is_previous_bigger_than_Z = False
    is_appended_next_symbol = False
    for index in range(len(row_alpha)):
        if len(row_alpha) == 1:
            if ord(row_alpha[index]) < 90 - step:
                next_row_alpha.append(chr(ord(row_alpha[index]) + step))
            else:
                next_char_ASCII_index = 65 + 9 - (90 - ord(row_alpha[index]))
                next_row_alpha.append(chr(next_char_ASCII_index))
                next_row_alpha.append('A')

        elif len(row_alpha) > 1:
            if is_appended_next_symbol:
                next_row_alpha.append(row_alpha[index])
                continue
            if is_previous_bigger_than_Z:
                if ord(row_alpha[index]) + 1 > 90 and index == len(row_alpha) - 1:
                    if row_alpha[index] == 'Z':
                        next_row_alpha.append('A')
                        next_row_alpha.append('A')
                        break
                    else:
                        next_char_ASCII_index = 65 + 9 - (90 - ord(row_alpha[index]))
                        next_row_alpha.append(chr(next_char_ASCII_index))
                        next_row_alpha.append('A')
                        break
                else:
                    next_row_alpha.append(chr(ord(row_alpha[index]) + 1))
                    is_appended_next_symbol = True
                    continue
            if ord(row_alpha[index]) > 90 - step:
                next_char_ASCII_index = 65 + 9 - (90 - ord(row_alpha[index]))
                next_row_alpha.append(chr(next_char_ASCII_index))
                is_previous_bigger_than_Z = True
                continue
            else:
                next_row_alpha.append(chr(ord(row_alpha[index]) + step))
                is_appended_next_symbol = True
    return ''.join(list(reversed(next_row_alpha)) + row_digit)


print("Starting Index of 'Aa1' for current Shear Wall: ", end='')
starting_index_for_Aa1 = input()
print("Starting Index of value for 'Aah' for current Shear Wall: ", end='')
ending_index_for_Aah = input()
print("Building levels for current Shear Wall: ", end='')
building_levels = int(input())
print("Valid starting index of row for head_size "
      "(Hint: Choose '1' if it's first shear wall, and 'previous row index+15' for each other): ", end='')
row_index_for_head_size = int(input())
# for each Shear Wall column indeces start from 'F'=6
column_index_for_head_size = 6

for level in range (building_levels):

    # find choosen from engineer head size for each level and calculate number of rebars in it
    head_size = find_head_size(row_index_for_head_size, column_index_for_head_size)
    head_rebars = calculate_rebars_in_head(head_size[0], head_size[1])
    print(head_size)
    print(head_rebars)
    column_index_for_head_size += 9
    if level == 0:
        step = 0
    else:
        step = 10

    first_row_index = find_required_index(starting_index_for_Aa1, step)
    end_row_index = find_required_index(ending_index_for_Aah, step)
    data_rows = []

    # read excel cell with required reinforcement for each level
    i=0
    for row in sheet[first_row_index: end_row_index]:
        row_as_list=str(row)
        dot_index=row_as_list.index('.')
        compare_symbol_index=row_as_list.index('>')
        row_as_list= row_as_list[dot_index + 1:compare_symbol_index]
        # data_columns = []
        for cell in row:
            # data_columns.append(cell.value)
        # data_rows.append(data_columns)
    # df = pd.Series(data_rows)
    # print(df)
            if i <= 1:
                result = calculate_Aa1_Aa2(cell.value, head_rebars)
                print(f"{head_rebars}N{result}")
                sheet[find_required_index(row_as_list, 4)].value = head_rebars
                sheet[find_required_index(row_as_list, 5)].value = result
            elif i == 2:
                result = calculate_AaV(cell.value)
                print(f"{result[0]} -> N{result[1]}/{result[2]}")
                sheet[find_required_index(row_as_list, 4)].value = result[0]
                sheet[find_required_index(row_as_list, 5)].value = result[1]
                sheet[find_required_index(row_as_list, 6)].value = '/' + str(result[2])
            else:
                result = calculate_AaH(cell.value)
                print(f"{result[0]} -> N{result[1]}/{result[2]}")
                sheet[find_required_index(row_as_list, 4)].value = result[0]
                sheet[find_required_index(row_as_list, 5)].value = result[1]
                sheet[find_required_index(row_as_list, 6)].value = '/' + str(result[2])
            workbook.save('Shear walls_TOWER.xlsx')
        i+=1
    # searching for new values of required reinforcement starts from the last found indices
    starting_index_for_Aa1 = first_row_index
    ending_index_for_Aah = end_row_index


