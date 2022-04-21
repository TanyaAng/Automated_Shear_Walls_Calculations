import openpyxl
from math import ceil
import json
from tkinter import messagebox

REBARS_DIAMETER = [8, 10, 12, 14, 16, 18, 20, 22, 25, 28, 32]
REBARS_CM2 = [0.503, 0.785, 1.131, 1.539, 2.011, 2.545, 3.142, 3.801, 4.909, 6.158, 8.042]
REBAR_WEIGHT = [0.395, 0.617, 0.888, 1.208, 1.998, 2.466, 2.984, 3.853, 4.834, 6.313, 7.990]


def find_head_size(row_indices, column_indices, sheet):
    column_size = []
    # for x_indices in range(row_indices, row_indices + 8):
    for y_indices in range(column_indices, column_indices + 10):
        if sheet.cell(row=row_indices, column=y_indices).value == 'глава - h/b':
            column_size.append(sheet.cell(row=row_indices, column=y_indices + 4).value)
            column_size.append(sheet.cell(row=row_indices, column=y_indices + 5).value)
            break
    return column_size


def calculate_rebars_in_head(height, width):
    number_rebars = 0
    if height == 10:
        number_rebars += 2
    elif height > 10:
        number_rebars += ceil((height - 5) / 15 + 1) * 2
    if width > 25:
        if width in [30, 35, 40, 45] and height == 10:
            number_rebars += 1
        elif width in [50, 60] and height == 10:
            number_rebars += 2
        elif width in [30, 35, 40, 45] and height > 10:
            number_rebars += 2
        else:
            number_rebars += ceil((width - 5) / 15 - 1) * 2
    return number_rebars


def calculate_Aa1_Aa2(required_reinforcement, number_of_bars):
    required_diameter = required_reinforcement / number_of_bars
    for index in range(len(REBARS_CM2)):
        if required_diameter <= 1.539:
            provided_diameter = 14
        else:
            if required_diameter <= REBARS_CM2[index]:
                provided_diameter = REBARS_DIAMETER[index]
                break
            else:
                provided_diameter = 'Change column size!!!'
    return provided_diameter


def calculate_AaH(required_reinforcement):
    steps = [20, 15, 12.5, 10]
    rebars_per_meter = [round(100 / x, 2) for x in steps]
    min_weight = 10000000
    index_of_min_rebar = 0
    index_of_min_step = 0
    for rebar in REBARS_CM2:
        needed_rebars_per_meter = required_reinforcement / rebar
        index = REBARS_CM2.index(rebar)
        current_rebar_weight = REBAR_WEIGHT[index]
        for i in range(len(rebars_per_meter)):
            if needed_rebars_per_meter > rebars_per_meter[i]:
                continue
            else:
                weight_of_current_mesh = rebars_per_meter[i] * current_rebar_weight
            if 0 < weight_of_current_mesh <= min_weight:
                min_weight = weight_of_current_mesh
                index_of_min_rebar = index
                index_of_min_step = i
    return rebars_per_meter[index_of_min_step], REBARS_DIAMETER[index_of_min_rebar], steps[index_of_min_step]


def calculate_AaV(required_reinforcement, head_width):
    wall_width = head_width
    min_percent_vertical_reinfocement = 0.002
    min_reinforcement = min_percent_vertical_reinfocement * 100 * wall_width / 2
    steps = [20, 15]
    rebars_per_meter = [round(100 / x, 2) for x in steps]
    min_weight = 10000000
    index_of_min_rebar = 0
    index_of_min_step = 0
    for rebar in REBARS_CM2:
        needed_rebars_per_meter = required_reinforcement / rebar
        index = REBARS_CM2.index(rebar)
        current_rebar_weight = REBAR_WEIGHT[index]
        for i in range(len(rebars_per_meter)):
            if needed_rebars_per_meter > rebars_per_meter[i]:
                continue
            else:
                weight_of_current_mesh = rebars_per_meter[i] * current_rebar_weight
            if 0 < weight_of_current_mesh <= min_weight and rebar * rebars_per_meter[i] > min_reinforcement:
                min_weight = weight_of_current_mesh
                index_of_min_rebar = index
                index_of_min_step = i
    return rebars_per_meter[index_of_min_step], REBARS_DIAMETER[index_of_min_rebar], steps[index_of_min_step]


# This function works only for row index with two leters (example:HJ...)
def find_required_index(starting_row, step):
    row_alpha = []
    row_digit = []
    for index in starting_row:
        if index.isalpha():
            row_alpha.append(index)
        elif index.isdigit():
            row_digit.append(index)
    row_alpha = list(reversed(row_alpha))
    next_row_alpha = []
    is_previous_bigger_than_Z = False
    is_appended_next_symbol = False
    for index in range(len(row_alpha)):
        if len(row_alpha) == 1:
            if ord(row_alpha[index]) < 90 - step:
                next_row_alpha.append(chr(ord(row_alpha[index]) + step))
            else:
                next_char_ASCII_index = 65 + (step - 1) - (90 - ord(row_alpha[index]))
                next_row_alpha.append(chr(next_char_ASCII_index))
                next_row_alpha.append('A')

        elif len(row_alpha) > 1:
            if is_appended_next_symbol:
                next_row_alpha.append(row_alpha[index])
                continue
            if is_previous_bigger_than_Z:
                if ord(row_alpha[index]) + 1 > 90 and index == len(row_alpha) - 1:
                    if row_alpha[index] == 'Z':
                        next_row_alpha.append('A')
                        next_row_alpha.append('A')
                        break
                    else:
                        next_char_ASCII_index = 65 + (step - 1) - (90 - ord(row_alpha[index]))
                        next_row_alpha.append(chr(next_char_ASCII_index))
                        next_row_alpha.append('A')
                        break
                else:
                    next_row_alpha.append(chr(ord(row_alpha[index]) + 1))
                    is_appended_next_symbol = True
                    continue
            if ord(row_alpha[index]) > 90 - step:
                next_char_ASCII_index = 65 + (step - 1) - (90 - ord(row_alpha[index]))
                next_row_alpha.append(chr(next_char_ASCII_index))
                is_previous_bigger_than_Z = True
                continue
            else:
                next_row_alpha.append(chr(ord(row_alpha[index]) + step))
                is_appended_next_symbol = True
    return ''.join(list(reversed(next_row_alpha)) + row_digit)


def find_current_shear_wall_index(number_of_shear_wall):
    next_starting_column = 'H8'
    next_ending_column = 'H11'
    if number_of_shear_wall >= 1:
        list_alpha = []
        list_digit = []
        for symbol in next_starting_column:
            if symbol.isdigit():
                list_digit.append(symbol)
            elif symbol.isalpha():
                list_alpha.append(symbol)
        string_of_digits = [str(el) for el in list_digit]
        int_of_digits = int(''.join(string_of_digits))
        next_starting_column = str(int_of_digits + 12 * (number_of_shear_wall))
        string_of_alpha = ''.join(list_alpha)
        next_starting_column = string_of_alpha + next_starting_column
        list_alpha = []
        list_digit = []
        for symbol in next_ending_column:
            if symbol.isdigit():
                list_digit.append(symbol)
            elif symbol.isalpha():
                list_alpha.append(symbol)
        string_of_digits = [str(el) for el in list_digit]
        int_of_digits = int(''.join(string_of_digits))
        next_ending_column = str(int_of_digits + 12 * (number_of_shear_wall))
        string_of_alpha = ''.join(list_alpha)
        next_ending_column = string_of_alpha + next_ending_column

    return next_starting_column, next_ending_column


def calculate_shear_wall(count_shear_walls, levels, file_path, sheet, workbook):
    number_of_shear_walls = count_shear_walls
    building_levels = levels
    row_index_for_head_size = 7
    for current_shear_wall in range(number_of_shear_walls):
        column_index_for_head_size = 6
        starting_index_for_current_wall = find_current_shear_wall_index(current_shear_wall)
        index_for_Aa1 = starting_index_for_current_wall[0]
        index_for_Aah = starting_index_for_current_wall[1]
        print(f"<<< SHEAR WALL {current_shear_wall + 1} >>>")
        for level in range(building_levels):
            print(f'Level:{level + 1}')
            # find choosen from engineer head size for each level and calculate number of rebars in it
            head_height, head_width = find_head_size(row_index_for_head_size, column_index_for_head_size, sheet)
            head_rebars = calculate_rebars_in_head(head_height, head_width)
            print(f"{head_rebars} number of rebars in head {head_height}, {head_width}")
            column_index_for_head_size += 10
            if level == 0:
                step = 0
            else:
                step = 10
            start_row_index = find_required_index(index_for_Aa1, step)
            end_row_index = find_required_index(index_for_Aah, step)
            # searching for new values of required reinforcement starts from the last found indices
            index_for_Aa1 = start_row_index
            index_for_Aah = end_row_index
            # read excel cell with required reinforcement for each level
            cell_number = 0
            for row in sheet[start_row_index: end_row_index]:
                current_row_as_string = str(row)
                dot_index = current_row_as_string.index('.')
                compare_symbol_index = current_row_as_string.index('>')
                current_row_as_string = current_row_as_string[dot_index + 1:compare_symbol_index]
                for cell in row:
                    # if cell.value == None:
                    #     cell.value = 0
                    if cell.value != None:
                        if cell_number == 0 or cell_number == 1:
                            result = calculate_Aa1_Aa2(cell.value, head_rebars)
                            print(f"{head_rebars}N{result}")
                            sheet[find_required_index(current_row_as_string, 4)].value = head_rebars
                            sheet[find_required_index(current_row_as_string, 5)].value = result
                        elif cell_number == 2:
                            result = calculate_AaV(cell.value, head_width)
                            print(f"{result[0]} -> N{result[1]}/{result[2]}")
                            sheet[find_required_index(current_row_as_string, 4)].value = result[0]
                            sheet[find_required_index(current_row_as_string, 5)].value = result[1]
                            sheet[find_required_index(current_row_as_string, 6)].value = '/' + str(result[2])
                        elif cell_number == 3:
                            result = calculate_AaH(cell.value)
                            print(f"{result[0]} -> N{result[1]}/{result[2]}")
                            sheet[find_required_index(current_row_as_string, 4)].value = result[0]
                            sheet[find_required_index(current_row_as_string, 5)].value = result[1]
                            sheet[find_required_index(current_row_as_string, 6)].value = '/' + str(result[2])
                cell_number += 1
        row_index_for_head_size += 12

    workbook.save(file_path)


# if __name__ == '__main__':
def main_function_calculate():
    with open('database.txt', 'r') as file:
        try:
            user_input = json.load(file)
            txt_path = user_input[0]
            file_path = user_input[1]
            sheet_name = user_input[2]
            number_shear_walls = int(user_input[3])
            storey_levels = int(user_input[4])

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook[sheet_name]
            calculate_shear_wall(number_shear_walls, storey_levels, file_path, sheet, workbook)
            print(user_input)
            messagebox.showinfo("Info", "Successful")
        except:
            messagebox.showwarning("Warning", "Invalid Input")
